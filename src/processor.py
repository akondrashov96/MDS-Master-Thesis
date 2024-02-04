import pandas as pd
import numpy as np
import os
import re
import locale
import pymorphy3
from dateutil import parser
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from copy import copy

from src.logger import logFile
from src.utilities import RussianParserInfo
from src.spellcheck import correct_errors, name_reconstruct, address_reconstruct


# Настройки
workdir = "data/"
locale.setlocale(locale.LC_ALL, 'ru_RU')
m = pymorphy3.MorphAnalyzer()


def correct_date_condition2(datestring: str) -> str:

    """
    Функция для проверки и корректировки по условию 2: 
    Графы «Поступление» и «Увольнение» пункта 14 даты должны содержать только цифры и точки.

    Функция проверяет соответствие даты формату ДД.ММ.ГГГГ или ММ.ГГГГ.
    Функция угадывает формат даты, если указана дата в читаемом формате и возвращает строку в формате ММ.ГГГГ.
    В противном случае функция оставляет в строке только числа и точки. 
    В случае, если указано "по настоящее время" - текст не меняется.

    Параметры:
    datestring : str
        Дата для форматирования

    Возвращает:
    string_out : str
        Строка с датой в формате ДД.ГГГГ

    """

    if datestring.lower() == "по настоящее время": return datestring
    
    try:

        dt = datetime.strptime(datestring, "%m.%Y")
        return(datestring)
    
    except ValueError as e:

        try:

            dt = datetime.strptime(datestring, "%d.%m.%Y")
            return(dt.strftime("%m.%Y"))
        
        except ValueError as e:

            try: 
                
                return(date_guesser_corrector(datestring, "%m.%Y"))
            
            except Exception as e:

                # нечитаемый формат даты
                pass

    return("".join(re.findall(pattern = "[0-9.]+", string = datestring)))


def only_cyrillic(textstring: str) -> str:

    """
    Функция для проверки и корректировки по условию 3.1: 
    1. Графа «Степень родства» пункта 15 должна содержать только буквы кириллицы

    Параметры:
    textstring : str
        Текст для форматирования

    Возвращает:
    string_out : str
        Строка, содержащая текст, состоящий только из символов кириллицы

    """

    return("".join(re.findall(pattern = "[а-яА-Я ]+", string = textstring)))


def last_surnames(textstring: str) -> str:

    """
    Функция для проверки и корректировки по условию 3.2: 
    2. В графе «Фамилия, имя и отчество» предыдущие фамилии (девичьи, изменённые) указываются в скобках. 
    Если у одного родственника несколько раз изменялась фамилия, то они указываются в скобках через запятую.

    Параметры:
    textstring : str
        Текст для форматирования

    Возвращает:
    string_out : str
        Строка, содержащая имя в соответсвующем формате

    """

    x = " ".join(re.findall(pattern = "[а-яА-Я- ]+", string = textstring)).split()

    if len(x) > 3:
        string_out = x[0] + " (" + ", ".join(x[1:-2]) + ") " + " ".join(x[-2:])
    else:
        string_out = " ".join(x)

    return string_out


def sort_address(adress_text: str) -> str:

    """
    Функция для проверки и корректировки по условию 4: 
    При заполнении адресов проживания и работы сначала необходимо указывать регион: республику, край, область.

    23.01.2024: DEPRECATED: Функция заменена на функцию address_reconstruct() из модуля spellcheck, 
    данная функцию использует нейросеть для определения адресов.

    Параметры:
    adress_text : str
        Адрес для форматирования

    Возвращает:
    adr_classes: numpy array of str
        Массив из классов элементов адреса

    string_out : str
        Строка, содержащая адрес в необходимом формате

    """

    # инициализация паттернов
    pattern_reg = "республика|респ\.|область|обл\.|край|кр\.|асср"
    pattern_subreg = "район|р-н"
    pattern_towncity = "г\.|гор\.|город|п\.|пос\.|поселок|гп|городское поселение|с\.|село"
    pattern_street = "ул\.|улица|б-р|бульвар|пр\.|проезд"
    pattern_house = "дом|д\."
    pattern_flat = "квартира|кв\."

    patterns = {"REGION": pattern_reg, 
                "SUBREGION": pattern_subreg, 
                "TOWNCITY": pattern_towncity, 
                "STREET": pattern_street, 
                "HOUSE": pattern_house,
                "FLAT": pattern_flat,
                "UNDEFINED": None}
    
    # создание словаря для сортировки элементов адреса
    sort_dict = {key: elem for elem, key in list(enumerate(patterns.keys()))}

    adr_tokens = adress_text.split(", ")
    adr_classes = np.full(len(adr_tokens), "UNDEFINED")

    # создание массива элементов адресов
    for i, adr_token in enumerate(adr_tokens):

        for key, pattern in patterns.items():

            if key == "UNDEFINED": continue

            # print(key)

            res = re.findall(pattern, adr_token.lower())

            if len(res) > 0:

                if adr_classes[i] != "UNDEFINED":
                    
                    adr_classes[i] += "|" + key

                else:

                    adr_classes[i] = key

    # print(adr_classes)
    # print(adr_tokens)

    # переформирование адреса
    string_out = ", ".join([x for _, x in sorted(zip(adr_classes, adr_tokens), key = lambda pair: sort_dict[pair[0]])])

    return string_out


def case_corrector(string_in: str, word: str, case: str = "nomn") -> str:

    """
    Функция для исправления падежа слова в предложении. 
    Делается через подстановку заполняемого выражения в строку и форматирование.

    Параметры:
    string_in : str
        Строка, в которой есть слово для форматирования
    word : str
        Слово, которое необходимо скорректировать (используется как паттерн)
    case : str
        Падеж, в который необходимо просклонять слово

    Возвращает:
    string_out : str
        Строка, содержащая слово в желаемом падеже
    """

    string_out = re.sub(word, "{replacement}", string_in)\
            .format(replacement = m.parse(word)[0].inflect({case}).word.title())
    
    return string_out


def date_guesser_corrector(datestring: str, target_format: str = "%Y, %d %B") -> str:

    """
    Функция для определения datetime из произвольного формата даты и конвертации полученного
    datetime в указанный формат.

    Параметры:
    datestring : str
        Строка, содержащая дату в произвольном формате (на русском)
    target_format : str
        Желаемый формат даты

    Возвращает:
    datestring_corrected : str
        Строка, содержащая дату в желаемом формате
    """

    # Создание паттернов для поиска форматов и частей даты
    date_pattern = "\%[a-zA-Z]"
    pattern_split = "[0-9а-яА-Я]+"

    formats = re.findall(date_pattern, target_format)

    # Угадывание даты из строки
    guessdate = parser.parse(datestring, 
                             parserinfo = RussianParserInfo())                         
    
    # если в искомом формате присутствует месяц, необходимо применить верное склонение
    if "%B" in formats: 
        
        guessdate_inner = guessdate.strftime("%d %B %Y")
        dateparts_inner = re.findall(pattern_split, guessdate_inner)

        datestring_corrected = case_corrector(guessdate.strftime(target_format), dateparts_inner[1], 'gent')
    
    else:

        datestring_corrected = guessdate.strftime(target_format)
    
    return datestring_corrected


def save_to_excel(df_list: list, filename: str):

    """
    Функция для записи списка таблиц в шаблон Excel. 
    На вход получает обработанный список из таблиц формы, список состоит из 6 элементов:
     - 1-2 таблицы - Лист 1
     - 3 таблица - Лист 2
     - 4 таблица - Лист 3
     - 5-6 таблицы - Лист 4
    
    Функция записывает данные таблицы в шаблон, хранящийся в директории templates
    При необходимости, производится добавление строк (с копированием форматирования ячеек)

    Функция ничего не возвращает, отрабатывает как процедура.
    """

    template_file = "templates/form4.template.xlsx"
    output_file = "data/processed/" + filename.replace(".xlsx", "") + "_check.xlsx"

    wb = load_workbook(template_file)

    # Запись первого листа
    ws = wb['Лист1']

    # Часть ФИО
    rows = dataframe_to_rows(df_list[0][["Ответ"]], index=False, header=False)

    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx+3, column=c_idx+1, value=value)

    # Часть ответа на вопросы
    rows = dataframe_to_rows(df_list[1][["Ответ"]], index=False, header=False)

    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx+7, column=c_idx+1, value=value)

    # Запись второго листа
    ws = wb['Лист2']

    rows = dataframe_to_rows(df_list[2], index=False, header=False)

    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx+3, column=c_idx, value=value)

    # Запись третьего листа
    ws = wb['Лист3']

    rows = dataframe_to_rows(df_list[3], index=False, header=False)

    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx+2, column=c_idx, value=value)

    # Запись четвертого листа
    ws = wb['Лист4']

    # определение количества строк в п. 17 - добавление строк при необходимости
    if len(df_list[5]) > 11:
        rows_to_add2 = len(df_list[5]) - 11
        ws.insert_rows(20, rows_to_add2)
    else:
        rows_to_add2 = 0

    # определение количества строк в п. 16 - добавление строк при необходимости
    if len(df_list[4]) > 4:
        rows_to_add1 = len(df_list[4]) - 4
        ws.insert_rows(7, rows_to_add1)
    else:
        rows_to_add1 = 0

    # сохранение стиля ячейки
    cellstyle = {"fill": copy(ws.cell(4, 1).fill), 
                "font": copy(ws.cell(4, 1).font), 
                "number_format": copy(ws.cell(4, 1).number_format), 
                "border": copy(ws.cell(4, 1).border)}

    # запись п. 16
    rows = dataframe_to_rows(df_list[4], index=False, header=False)

    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):

            # print(r_idx)

            ws.cell(row=r_idx+3, column=c_idx).value = value
            ws.cell(row=r_idx+3, column=c_idx).fill = (cellstyle['fill'])
            ws.cell(row=r_idx+3, column=c_idx).font = (cellstyle['font'])
            ws.cell(row=r_idx+3, column=c_idx).number_format = (cellstyle['number_format'])
            ws.cell(row=r_idx+3, column=c_idx).border = (cellstyle['border'])

    # запись п. 17
    rows = dataframe_to_rows(df_list[5], index=False, header=False)

    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx+9+rows_to_add1, column=c_idx).value = value
            ws.cell(row=r_idx+9+rows_to_add1, column=c_idx).fill = (cellstyle['fill'])
            ws.cell(row=r_idx+9+rows_to_add1, column=c_idx).font = (cellstyle['font'])
            ws.cell(row=r_idx+9+rows_to_add1, column=c_idx).number_format = (cellstyle['number_format'])
            ws.cell(row=r_idx+9+rows_to_add1, column=c_idx).border = (cellstyle['border'])

    wb.save(output_file)
    wb.close()


def file_processor(filename: str, workdir: str = workdir, logfile: str = "", verbose: bool = False):

    """
    Чтение и предобработка данных для каждого листа анкеты. 

    На листах осуществляется следующая обработка:
    
    Лист 1:

        Содержимое листа разбивается на 2 части - ФИО и ответы на пп. 2-13

        В ячейках ответов на пп. 2-13 осуществляется проверка орфографии
        
        Проверяется условие 1: В пункте 3 год рождения указывается только цифрами, число – двумя цифрами
        
    Лист 2:
        
        Проверяется условие 2: Графы «Поступление» и «Увольнение» пункта 14 даты должны содержать только цифры и точки.

        В колонках 3-4 
            - Должность с указанием наименования организации и 
            - Адрес организации (фактический, юридический, в т.ч. за границей)
        осуществляется проверка орфографии.

        В колонке Адрес организации (фактический, юридический, в т.ч. за границей) проверяется условие 4:
            При заполнении адресов проживания и работы сначала необходимо указывать регион: республику, край, область.

    Лист 3:

        Проверяются условия 3.1 и 3.2: 
            1. Графа «Степень родства» пункта 15 должна содержать только буквы кириллицы.
            2. В графе «Фамилия, имя и отчество» предыдущие фамилии (девичьи, изменённые) указываются в скобках. 
            Если у одного родственника несколько раз изменялась фамилия, то они указываются в скобках через запятую.

        В колонках 3-5
            - Число, месяц, год и место рождения, гражданство, 
            - Место работы, должность, 
            - Адрес места жительства, а также откуда и когда прибыл ***
        Производится проверка орфографии.

        В колонке Адрес места жительства, а также откуда и когда прибыл *** проверяется условие 4:
            При заполнении адресов проживания и работы сначала необходимо указывать регион: республику, край, область.

    Лист 4:

        Содержимое листа разбивается на 2 части - ответы на п. 16 и ответ на п. 17

        Проверяются условия 3.1 и 3.2: 
            1. Графа «Степень родства» пункта 16 должна содержать только буквы кириллицы.
            2. В графе «Фамилия, имя и отчество» пункта 16 предыдущие фамилии (девичьи, изменённые) указываются в скобках. 
            Если у одного родственника несколько раз изменялась фамилия, то они указываются в скобках через запятую.

        В таблице п. 17:

        Проверяется условие 2: 
            Графы Периода проживания пункта 17 даты должны содержать только цифры и точки.

        Производится проверка орфографии в колонке адреса.

        Проверяется условие 4: 
            При заполнении адресов проживания и работы сначала необходимо указывать регион: республику, край, область.
    
    Параметры:
    filename : str
        Имя файла
    workdir : str
        Рабочая директория

    Возвращает:
    pd_list : str of pd.DataFrame
        Список из таблиц для последующем сохранении в Excel
    """

    if verbose: print(f"Обработка документа {filename}")

    # Начать логирование: лог 1
    if logfile == "" and len(os.listdir("logs")) != 0:

        logfile = os.listdir("logs")[-1]

    log = logFile(mode = "a", filename = "logs/"+logfile, operation = "Обработка")

    data_sheets = [pd.read_excel(workdir + "raw/" + filename,
                                 sheet_name=sheet_index) for sheet_index in range(4)]
    
    # Лог 2
    msg = f"{filename}: Листы считаны"
    log.write_log(msg)
    if verbose: print(msg)

    
    # Лист 1: обработка

    # Лог 3
    msg = f"{filename}: Обработка первого листа"
    log.write_log(msg)
    if verbose: print(msg)

    # Переименование колонок и удаление пустот
    data_sheets[0].columns = ["Вопрос", "Ответ"]
    data_sheets[0] = data_sheets[0][2:].dropna(subset = ["Вопрос"]).reset_index(drop = True)

    data_sheets_0_1 = data_sheets[0][:3].reset_index(drop = True)
    data_sheets_0_2 = data_sheets[0][3:].reset_index(drop = True) 

    # Проверка орфографии
    if verbose: print("Проверка орфографии... Лист 1")
    data_sheets_0_2["Ответ"] = data_sheets_0_2["Ответ"].map(correct_errors)
    
    # Лог 4
    msg = f"{filename}: Лист 1: орфография проверена"
    log.write_log(msg)
    if verbose: print(msg)

    # Проверка условия 1: В пункте 3 год рождения указывается только цифрами, число – двумя цифрами 
    date_place_parts = data_sheets_0_2['Ответ'][1].split(", ")

    # Проверка на вхождение паттерна
    regex_matcher = re.search(pattern = r"\d{4},\s\d{2}\s\w+", 
                              string = date_place_parts[0] + ", " + date_place_parts[1])
    
    if regex_matcher is None:

        # Вычленение даты произвольного формата и переформативание в ГГГГ, ДД ММММ
        try:

            date_corrected = date_guesser_corrector(date_place_parts[0] + ", " + date_place_parts[1])

            date_place_corrected = " ".join([date_corrected, 
                                             *date_place_parts[2:]])
            
            data_sheets_0_2['Ответ'][1] = date_place_corrected

        except Exception as e:
            
            # Лог error_1
            msg = f"{filename}: Лист 1: ошибка в формате даты"
            log.write_log(msg, content = "ERR")
            log.write_log(str(e), content = "ERR")

    else:

        # Нужно ли исправить падеж в дате при вхождении паттерна?
        pass

    # Лог 5
    msg = f"{filename}: Лист 1: Условие 1 исправлено"
    log.write_log(msg)
    if verbose: print(msg)

    if verbose: print(f"Обработка листа 1 завершена")


    # Лист 2: обработка

    # Лог 6
    msg = f"{filename}: Обработка второго листа"
    log.write_log(msg)
    if verbose: print(msg)

    data_sheets[1].columns = ["Месяц и год поступления", 
                              "Месяц и год увольнения",
                              "Должность с указанием наименования организации",
                              "Адрес организации"]

    data_sheets[1] = data_sheets[1][2:].dropna().reset_index(drop = True)

    # Проверка условия 2: Графы «Поступление» и «Увольнение» пункта 14 даты должны содержать только цифры и точки.
    data_sheets[1]["Месяц и год поступления"] = data_sheets[1]["Месяц и год поступления"].map(correct_date_condition2)
    data_sheets[1]["Месяц и год увольнения"] = data_sheets[1]["Месяц и год увольнения"].map(correct_date_condition2)

    # Лог 7
    msg = f"{filename}: Лист 2: Условие 2 исправлено"
    log.write_log(msg)
    if verbose: print(msg)

    # Проверка орфографии
    if verbose: print("Проверка орфографии Лист 2...")
    data_sheets[1]["Должность с указанием наименования организации"] = \
        data_sheets[1]["Должность с указанием наименования организации"].map(correct_errors)
    data_sheets[1]["Адрес организации"] = data_sheets[1]["Адрес организации"].map(correct_errors)
    
    # Лог 8
    msg = f"{filename}: Лист 2: орфография проверена"
    log.write_log(msg)
    if verbose: print(msg)

    # Проверка условия 4: При заполнении адресов проживания и 
    # работы сначала необходимо указывать регион: республику, край, область.
    if verbose: print(f"Проверка условия 4")

    # Переупорядочивание элементов адреса
    data_sheets[1]["Адрес организации"] = data_sheets[1]["Адрес организации"].map(address_reconstruct)

    # Лог 9
    msg = f"{filename}: Лист 2: Условие 4 исправлено"
    log.write_log(msg)
    if verbose: print(msg)

    if verbose: print(f"Обработка листа 2 завершена")
    

    # Лист 3: обработка

    # Лог 10
    msg = f"{filename}: Обработка третьего листа"
    log.write_log(msg)
    if verbose: print(msg)

    # Замена имен колонок
    data_sheets[2].columns = ["Степень родства", 
                              "Фамилия, имя и отчество",
                              "Число, месяц, год и место рождения, гражданство",
                              "Место работы, должность",
                              "Адрес места жительства"]

    data_sheets[2] = data_sheets[2][1:].dropna().reset_index(drop = True)

    # Проверка условия 3.1: Графа «Степень родства» пункта 15 должна содержать только буквы кириллицы.
    if verbose: print(f"Проверка условия 3.1")

    data_sheets[2]['Степень родства'] = data_sheets[2]['Степень родства'].map(only_cyrillic)

    # Лог 11
    msg = f"{filename}: Лист 3: Условие 3.1 исправлено"
    log.write_log(msg)
    if verbose: print(msg)

    # Проверка условия 3.2: 
    # В графе «Фамилия, имя и отчество» предыдущие фамилии (девичьи, изменённые) указываются в скобках. 
    # Если у одного родственника несколько раз изменялась фамилия, то они указываются в скобках через запятую.
    if verbose: print(f"Проверка условия 3.2 (используется NER для имен)")

    # TO DO
    data_sheets[2]["Фамилия, имя и отчество"] = data_sheets[2]["Фамилия, имя и отчество"].map(name_reconstruct)

    # Лог 12
    msg = f"{filename}: Лист 3: Условие 3.2 исправлено"
    log.write_log(msg)
    if verbose: print(msg)

    # Проверка орфографии
    if verbose: print("Проверка орфографии Лист 3...")
    data_sheets[2]["Число, месяц, год и место рождения, гражданство"] = \
        data_sheets[2]["Число, месяц, год и место рождения, гражданство"].map(correct_errors)
    data_sheets[2]["Место работы, должность"] = data_sheets[2]["Место работы, должность"].map(correct_errors)
    data_sheets[2]["Адрес места жительства"] = data_sheets[2]["Адрес места жительства"].map(correct_errors)
    
    # Лог 13
    msg = f"{filename}: Лист 3: орфография проверена"
    log.write_log(msg)
    if verbose: print(msg)

    # Проверка условия 4: При заполнении адресов проживания и 
    # работы сначала необходимо указывать регион: республику, край, область.
    if verbose: print(f"Проверка условия 4 (используется NER для имен)")

    # Переупорядочивание элементов адреса (TO DO)
    data_sheets[2]["Адрес места жительства"] = data_sheets[2]["Адрес места жительства"].map(address_reconstruct)

    # Лог 14
    msg = f"{filename}: Лист 3: Условие 4 исправлено"
    log.write_log(msg)
    if verbose: print(msg)

    if verbose: print(f"Обработка листа 3 завершена")


    # Лист 4: обработка

    # Лог 15
    msg = f"{filename}: Обработка четвертого листа"
    log.write_log(msg)
    if verbose: print(msg)
    
    data_sheets_3_1 = data_sheets[3][2:5].dropna(how="all")

    data_sheets_3_1.columns = ["Степень родства", 
                               "Фамилия, имя и отчество",
                               "Где проживает и период проживания за границей"]

    stop_idx = data_sheets[3][data_sheets[3].iloc[:, 0].str.contains("Дополнительные сведения", na=False)].index[0]
    data_sheets_3_2 = data_sheets[3][7:stop_idx].dropna(how="all")

    data_sheets_3_2.columns = ["Период проживания начало",
                               "Период проживания конец",
                               "Адрес проживания и регистрации"]
    
    # Проверка условия 3.1: Графа «Степень родства» пункта 16 должна содержать только буквы кириллицы.
    if verbose: print(f"Проверка условия 3.1")

    data_sheets_3_1['Степень родства'] = data_sheets_3_1['Степень родства'].map(only_cyrillic)

    # Лог 16
    msg = f"{filename}: Лист 4: Условие 3.1 исправлено"
    log.write_log(msg)
    if verbose: print(msg)

    # Проверка условия 3.2: 
    # В графе «Фамилия, имя и отчество» предыдущие фамилии (девичьи, изменённые) указываются в скобках. 
    # Если у одного родственника несколько раз изменялась фамилия, то они указываются в скобках через запятую.
    if verbose: print(f"Проверка условия 3.2")

    # TO DO
    data_sheets_3_1["Фамилия, имя и отчество"] = data_sheets_3_1["Фамилия, имя и отчество"].map(name_reconstruct)

    # Лог 17
    msg = f"{filename}: Лист 4: Условие 3.2 исправлено"
    log.write_log(msg)
    if verbose: print(msg)

    # Проверка условия 2: Графы Периода проживания пункта 17 даты должны содержать только цифры и точки.
    data_sheets_3_2["Период проживания начало"] = data_sheets_3_2["Период проживания начало"].map(correct_date_condition2)
    data_sheets_3_2["Период проживания конец"] = data_sheets_3_2["Период проживания конец"].map(correct_date_condition2)

    # Лог 18
    msg = f"{filename}: Лист 4: Условие 2 исправлено"
    log.write_log(msg)
    if verbose: print(msg)

    # Проверка орфографии
    if verbose: print("Проверка орфографии Лист 4...")
    data_sheets_3_2["Адрес проживания и регистрации"] = data_sheets_3_2["Адрес проживания и регистрации"].map(correct_errors)
    
    # Лог 19
    msg = f"{filename}: Лист 4: орфография проверена"
    log.write_log(msg)
    if verbose: print(msg)

    # Проверка условия 4: При заполнении адресов проживания и 
    # работы сначала необходимо указывать регион: республику, край, область.
    if verbose: print(f"Проверка условия 4")

    # Переупорядочивание элементов адреса
    data_sheets_3_2["Адрес проживания и регистрации"] = data_sheets_3_2["Адрес проживания и регистрации"].map(address_reconstruct)

    # Лог 20
    msg = f"{filename}: Лист 4: Условие 4 исправлено"
    log.write_log(msg)
    if verbose: print(msg)

    if verbose: print(f"Обработка листа 4 завершена")

    # объединим обработанные таблицы в список
    pd_list = [data_sheets_0_1, data_sheets_0_2, data_sheets[1], data_sheets[2], data_sheets_3_1, data_sheets_3_2]

    save_to_excel(pd_list, filename)